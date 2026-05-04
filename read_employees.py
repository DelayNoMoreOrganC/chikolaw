#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""读取员工花名册Excel文件"""

from openpyxl import load_workbook
import json

excel_path = r"C:\Users\Administrator\Desktop\律所共享文件夹\投标参考资料\至高员工花名册（20260311更新）上报用.xlsx"

try:
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb.active

    print(f"Sheet名称: {sheet.title}")
    print(f"行数: {sheet.max_row}, 列数: {sheet.max_column}")

    # 尝试找到真正的表头行（包含"姓名"或"身份证"的行）
    header_row = 1
    for row in range(1, min(10, sheet.max_row + 1)):
        for col in range(1, sheet.max_column + 1):
            val = str(sheet.cell(row, col).value or "")
            if "姓名" in val or "身份证" in val:
                header_row = row
                break
        if header_row > 1:
            break

    print(f"\n表头行: 第{header_row}行")
    headers = []
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(header_row, col).value
        headers.append(cell_value)
        print(f"  列{col}: {cell_value}")

    print("\n前3行数据:")
    employees = []
    for row in range(header_row + 1, min(header_row + 4, sheet.max_row + 1)):
        employee = {}
        for col in range(1, sheet.max_column + 1):
            header = headers[col-1] if col <= len(headers) else f"列{col}"
            value = sheet.cell(row, col).value
            employee[header] = value
        employees.append(employee)
        print(f"\n第{row}行:")
        for k, v in employee.items():
            if v is not None:
                print(f"  {k}: {v}")

    # 保存所有数据
    all_employees = []
    for row in range(header_row + 1, sheet.max_row + 1):
        employee = {}
        for col in range(1, sheet.max_column + 1):
            header = headers[col-1] if col <= len(headers) else f"列{col}"
            value = sheet.cell(row, col).value
            employee[header] = value
        if any(v is not None for v in employee.values()):  # 跳过空行
            all_employees.append(employee)

    with open(r"D:\ZGAI\employees_data.json", "w", encoding="utf-8") as f:
        json.dump(all_employees, f, ensure_ascii=False, indent=2, default=str)

    print(f"\n[OK] 共读取 {len(all_employees)} 名员工，数据已保存到 employees_data.json")

except Exception as e:
    print(f"[ERROR] 错误: {e}")
    import traceback
    traceback.print_exc()

#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from docx import Document
from openpyxl import load_workbook
import json

# 读取Word文档
doc = Document(r"C:\Users\Administrator\Desktop\律所共享文件夹\系统行政管理要求.docx")
print("=" * 80)
print("系统行政管理要求.docx")
print("=" * 80)
for i, para in enumerate(doc.paragraphs):
    if para.text.strip():
        print(f"{i+1:3d}. {para.text}")

print("\n\n")

# 读取Excel文档
wb = load_workbook(r"C:\Users\Administrator\Desktop\律所共享文件夹\系统问题.xlsx")
print("=" * 80)
print("系统问题.xlsx")
print("=" * 80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n【工作表: {sheet_name}】")
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            print("  |  ".join(str(cell) if cell is not None else "" for cell in row))

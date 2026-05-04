#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
sys.stdout.reconfigure(encoding='utf-8')

from docx import Document
from openpyxl import load_workbook
import json

# 读取Word文档
doc = Document(r"C:\Users\Administrator\Desktop\律所共享文件夹\系统行政管理要求.docx")
print("=" * 80)
print("系统行政管理要求.docx")
print("=" * 80)
for i, para in enumerate(doc.paragraphs):
    if para.text.strip():
        print(f"{i+1:3d}. {para.text}")

print("\n\n")

# 读取Excel文档
wb = load_workbook(r"C:\Users\Administrator\Desktop\律所共享文件夹\系统问题.xlsx")
print("=" * 80)
print("系统问题.xlsx - 新建案件表单字段")
print("=" * 80)

ws = wb.active
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        print(f"表头: {row}")
        continue
    if any(cell is not None for cell in row):
        # 只打印非空行
        cells_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
        if cells_str.strip():
            print(f"行{i}: {cells_str}")
            if i > 30:  # 只打印前30行避免过长
                print("... (省略后续行)")
                break
